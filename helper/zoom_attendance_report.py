import csv
from datetime import datetime
from pathlib import Path
import logging
import string
import re

from fuzzywuzzy import process
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from .csv_parser import BaseCsvParser
try:
    from .manual_zoom_matches import MANUAL_FIXES
except ImportError:
    def MANUAL_FIXES(name):
        """
        For students with silly names the algorithm cannot recognize, write
        this function in the file above to make individual corrections.
        ./manual_zoom_matches is in the .gitignore because it will contain
        student names.
        """
        return name
from .helper import Helper

helper = Helper.read_cache()


class Meeting:
    """
    Single zoom meeting. Opens and reads a CSV file, as downloaded from
    zoom.

    Attrs:
        - rows
        - topic
        - datetime

    """

    def __init__(self, path: Path):
        self.path = path
        self.attendees = []
        self.datetime = None
        self.duration = None
        self.topic = None
        self.total_participants = None
        self.open_report()

    def __repr__(self):
        outstr = (
            f'<helper.zoom_attendance_report.Meeting; {self.topic} at '
            f'{self.datetime.isoformat()}>'
        )
        return outstr

    def __str__(self):
        return self.path.stem

    def __eq__(self, other):
        # input validation
        if not isinstance(other, Meeting):
            try:
                outstr = (
                    'Unsupported operand; cannot compare type "Meeting" to '
                    + type(other)
                )
                raise ValueError(outstr)
            finally:
                raise ValueError('Unsupported operand')
        self_identifier = self.topic + self.datetime.isoformat()
        other_identifier = other.topic + other.datetime.isoformat()
        if self_identifier == other_identifier:
            return True
        return False

    def __gt__(self, other):
        # input validation
        if not isinstance(other, Meeting):
            try:
                outstr = (
                    'Unsupported operand; cannot compare type "Meeting" to '
                    + type(other)
                )
                raise ValueError(outstr)
            finally:
                raise ValueError('Unsupported operand')
        return len(self.attendees) > len(other.attendees)


    def open_report(self):
        """
        Opens, reads, and parses zoom report such that IO will not need to
        be performed again.

        assigns attributes:
        self.topic
        self.grade_level
        self.datetime
        """
        try:
            self.grade_level = int(self.path.name[0])
        except ValueError:
            raise NotImplementedError(
                'Grade level is not the first character of the report name.'
            )
        with open(self.path, 'r') as csvfile:
            rows = [r for r in csv.reader(csvfile)]
        if rows[2]:
            raise Exception(
                'Zoom report must contain meeting information. The '
                f'Report at {self.path} does not appear to contain '
                'meeting information.'
            )
        self.duration = int(rows[1][5])
        self.topic = rows[1][1]
        time_str = rows[1][2]
        (
            month,
            day,
            year,
            hour,
            minute,
            *rest
        ) = [int(i) for i in re.split(r'/| |:', time_str) if i.isdigit()]
        if 'pm' in time_str.lower():
            hour += 12
        self.datetime = datetime(year, month, day, hour, minute)
        for st in helper.students.values():
            st.zoom_attendance_record = {}
        for row in rows[4:]:
            duration = row[2]
            st = self.match_student(row[0])
            if not st:
                continue
            st.zoom_attendance_record.setdefault(
                (self.topic + ';' + self.datetime.isoformat()),
                duration
            )
            self.attendees.append(st)

    def match_student(self, name):
        """
        Wrapper method that ties together the helper class's find_student
        method, the try_matching_student... method in this class, and some
        custom matches for students with really weird zoom names. Here is also
        where the student's raw zoom name is sanitized (punctuation removed).
        """
        # reference manual fixes, an optional import
        name = MANUAL_FIXES(name)

        # Rough cleaning
        # some use dot to delimit first / last name
        name.replace('.', ' ')
        # remove punctuation
        for char in string.punctuation:
            name.replace(char, '')
        # popular emoticon character
        name.replace('ω', '')
        try:
            return helper.find_nearest_match(name, auto_yes=True)
        except Warning:
            pass
        st = self.try_matching_student_within_grade(name)
        if st:
            return st
        for part in [i for i in re.split(' |.', name) if i]:
            st = self.try_matching_student_within_grade(part)
            if st:
                return st

    def try_matching_student_within_grade(self, student_name):
        """
        Unlike in the helper student matching function, this class is aware of
        the grade level of the student it is trying to match. If the helper
        method returns None, this fallback method tries to identify the
        student through process of elimination within their own grade. This
        helps match more students who only provide their first name.
        """
        first_name_match = process.extract(
            student_name,
            [
                s.first_name for s in helper.students.values()
                if s.grade_level == self.grade_level
            ],
            limit=3
        )
        if first_name_match[0][1] > 90:
            # we have a potential match! Let's see if it's truly a match
            # first, re-extract the student object from all students...

            # If the best two matches are the same name, there are two or more
            # students in the grade level with that name. It is therefore
            # impossible to perform a perfect match against only the first name
            if first_name_match[1][0] == first_name_match[0][0]:
                logging.debug(
                    f'Cannot proceed with {student_name}. More than one '
                    f'student in the {self.grade_level}th grade has the first '
                    f'name {first_name_match[0][0]}.'
                )

            # If the first two matches are not the same, that means the first
            # name is unique, and we can make a match within the grade level.
            for n, s in helper.students.items():
                if n.split(' ')[0] == first_name_match[0][0]:
                    logging.debug(
                        'Successful grade level match for '
                        + helper.students[n].name
                    )
                    return helper.students[n]


class MeetingSet:
    """

    # General Usage

    A directory full of meeting reports can be carelessly tossed into
    __init__. This class dynamically groups meetings by attendees. For example,
    if you use the same meeting topic (i.e. "Health") to meet with different
    groups of homerooms throughout the week, this class will observe the union
    of the set of attendees at each meeting instance, and if there is a
    significant intersection, those meetings are grouped.

    # Resultant Data Structure

    By default, meeting groupings will be accessible as a list (self.groups).
    Each item in the list will itself be a chronologically sorted list of
    Meeting instances. However, group_map may be passed to __init__ to produce
    a more descriptive data structure.

    # group_map: dict NOT CURRENTLY SUPPORTED

    The presumption is that there is no reliable way to know the full name of
    a meeting. For example, the meeting topic might be, "Health," but whoose
    homeroom is it? There's no way to know from here. group_map is a mapping
    of filenames to correct, fully-descriptive meeting names. An example
    group_map might look like:

    {
        '6th Grade Health 9-24-2020 10:13 am.csv': 'Health; Mrs. Smith's Homeroom',
        '6th Grade Health 9-25-2020 10:13 am.csv': 'Health; Mrs. Jones's Homeroom',
        'Garbled zoom report filename': 'A Group Label Useful to You'
    }

    Effectively, this "tags" one instance of a unique group with a particular
    name. These similar meeting instances are grouped anyway by default – that's
    basically the purpose of this class, but by tagging a single instance, you
    can get back a dict where these groups are named. Of course, it's also
    possible to just take the groupings from self.groups and assign names
    afterward yourself.

    # trust_topics: bool

    As I mentioned before, the presumption is that there is no reliable way to
    know the full name of a meeting. Often, teachers' meeting "topics," don't
    directly correspond to the names of the groups they are meeting with.
    If the meeting names are trustworthy, however, the group name will be the
    same as the meeting topic, and the same group_dict attribute as above will
    exist.

    """

    def __init__(self, dir_path: Path, group_map=None, trust_topics=False):
        self.dir_path = dir_path
        self.group_map = group_map
        if group_map:
            self.group_dict = {}
        self.trust_topics = trust_topics
        self.groups = []
        self.TOTAL_TO_UNION_RATIO_ADJUSTMENT = 0.75

    def process(self):
        """
        Called by __init__; produces data structure
        """




        # generate group map from filenames if needed

        # if (not self.group_map and self.trust_topics):
        #     self.generate_group_map_from_topics()





        # init group map; items in group map are pointers to nested lists in
        # self.groups

        # if self.group_map:
        #     for k, v in self.group_map.items():
        #         path = Path(self.dir_path, k)
        #         self.groups.append(li := [Meeting(path)])
        #         self.group_dict[v] = li





        # append all meetings to groupings in self.groups
        for meeting in self.iter_csvs():
            match = self.match_meeting_with_group_by_union(meeting)
            if match:
                if self.group_map:
                    if meeting.path.name in self.group_map:
                        # if the current meeting is the one the user associated
                        # with a specific name, put this Meeting object into the
                        # group_dict, which will act as a tag to flush out the
                        # group dict later
                        self.group_dict[self.group_map[meeting.path.name]] = [meeting]
                # append to group of similar meetings
                match.append(meeting)
            else:
                # create a new list of meetings if there is no match
                self.groups.append([meeting])





        # At this point, self.group_map, if it exists, only has a single tag
        # item which creates a link to a larger grouping in self.groups.
        # fill_group_dict uses that tag to find the matching group in
        # self.groups and fill out the rest of the group dict

        # if self.group_map:
        #     self.fill_group_dict()


        # TODO: deal with inconsistent attendance case:
        """


        AN IMPORTANT SIDE NOTE
        I don't know why, but commenting out the group_dict crap did improve
        the grouping result, so something up here is affecting the grouping
        algorithm negatively




        Union grouping breaks down when attendance between groups is sporradic.
        If the attendance rate for any meeting is below ~80%, The diff between
        the attendees of that partially-attended meeting and the full group
        will kick that meeting into a separate group.

        One way to deal with this may be to consider mutual exclusion.
        For the average teacher, attendees only come to one type of meeting.
        If we know that meeting attendees are mutually exclusive for any given
        meeting type, we can be much more presumptuous in creating groupings.

        Idea:
        If the meeting misses the union threshold by ~20%, and len(attendees)
        of the meeting is less than the max() (i.e.: this was a pooorly attended
        meeting), add it to the group anyway if the students in the current
        meeting are almost a perfect subset of the greater group.
        """


    def fill_group_dict(self):
        """
        Wherever a match was found for the user-provided label, that single
        meeting object was put into the group_dict. Now, we can use that meeting
        to find the group within self.groups and fill the meeitng with the whole
        group.
        """
        for grouping in self.groups:
            for meeting in grouping:
                label = self.group_map.get(meeting.path.name)
                if label:
                    self.group_dict[label] = grouping
                    break


    def match_meeting_with_group_by_union(self, meeting: Meeting):
        """
        Given a list of students, calculate the union between that list, and
        all the other lists of students in meetings previously provided.

        Return the group whose union against the provided list is less than the
        length of the lists combined, indicating that these are two instances
        of the same group of students meeting.
        """
        for group in self.groups:
            # using the biggest meeting (in terms of num of attendees)
            # so far for consistency
            past_meeting = max(group)
            pm__attendees = {s.name for s in past_meeting.attendees}
            cm__attendees = {s.name for s in meeting.attendees}
            union = len(pm__attendees | cm__attendees)
            total = len(pm__attendees) + len(cm__attendees)
            # Where groups are the same and attendance is perfect,
            # len(total) / 2 == len(union)
            # Considering imperfect attendance, the group will be considered
            # match if the union is less than 85% of the total.
            # This CONSTANT is definied in __init__
            if total * self.TOTAL_TO_UNION_RATIO_ADJUSTMENT > union:
                print(f'{meeting.__str__()} matches with {past_meeting.__str__()}')
                return group
        return []

    def generate_group_map_from_topics(self):
        """
        If trust_topics is true and group_map is None, this function
        generates a group map from the topics. This function will raise an
        exception if it should not have been called.
        """
        # validation
        try:
            assert not self.group_map
            assert self.trust_topics
        except AssertionError:
            raise Exception(
                'Preconditions for generate_group_map_from_topic were not '
                f'met.\ngroup_map:\t{self.group_map}\n\ntrust_topics:\t'
                + self.trust_topics
            )
        self.group_map = {}
        for path in self.dir_path.iterdir():
            meeting = Meeting(path)
            self.group_map.setdefault(path.name, meeting.topic)
        self.group_dict = {}
        return self.group_map

    def rename_csv_files(self):
        """
        Provide verbose names to csv file.
        """
        for report_path in self.dir_path.iterdir():
            with open(report_path, 'r') as csvf:
                rows = [r for r in csv.reader(csvf)]
            topic = rows[1][1]
            start_time = rows[1][2]
            date = start_time.split(' ')[0].replace('/', '-')
            new_name = topic + ' ' + date + '.csv'
            report_path.rename(Path(report_path.parent, new_name))

    def iter_csvs(self):
        """
        Skip anything that isn't a real csv file. Calls self.open_report(),
        which loads the report into memory as attributes of self; there is no
        need to yield anything, but it yields the path to the current csv
        for convenience.
        """
        for i in self.dir_path.iterdir():
            if i.name[-4:] != '.csv':
                continue
            if i.name[0] == "~" or i.name[0] == ".":
                continue
            yield Meeting(i)


class ExcelWriter:
    """
    Writes a MeetingSet into an excel report

    Optionally, pass attendance_duration_thresholds: dict as a kwarg. It should
    look like this:

        {
            'color: str': 'minimum minutes of attendance: int',
        }

        For example:

        {
            'red': 5,
            'yellow': 20,
            'green': 30
        }

        You must provide these colors:
            - 'red'
            - 'yellow'
            - 'green'
    """

    def __init__(self, meetings: MeetingSet, *a, **kw):
        super().__init__(*a, **kw)
        self.meetings = meetings
        self.thresholds = kw.get('attendance_duration_thresholds')
        self.validate_or_generate_thresholds()
        # init workbook
        self.WB_OUT = Workbook()
        self.WB_OUT.remove(self.WB_OUT.active)
        self.styles = {
            'h1': Font(size=30, bold=True, name='Cambria'),
            'h2': Font(size=16, name='Calibri')
        }

        def fill(colorcode):
            return PatternFill(
                fill_type='solid',
                start_color=colorcode,
                end_color=colorcode
            )

        self.fills = {
            'green': fill('18fc03'),
            'yellow': fill('fcf403'),
            'red': fill('fc0303')
        }

    def validate_or_generate_thresholds(self):
        if self.thresholds:
            for color in ['red', 'yellow', 'green']:
                assert color in self.thresholds
            assert (
                self.thresholds['red']
                < self.thresholds['yellow']
                < self.thresholds['green']
            )
            for i in self.thresholds.values():
                assert isinstance(i, int)
        else:
            # assign defaults
            self.thresholds = {}
            self.thresholds['red'] = 0
            self.thresholds['yellow'] = 15
            self.thresholds['green'] = 30

    def generate_report(self, destination: Path, thresholds=None):
        """
        Master Sheet:
            The master sheet contains a summary of all data in one place.
            The attendance log for each group is laid out from top to bottom.

        Group Sheets:
            Each group gets its own sheet for its attendance.

        Highlights Sheet(s):
            Gives some helpful highlights:
                - Students in the top and bottom 10th percentile of attendance
                    across all groups.
                - Group with the best & worst attendance (this is only measured
                    against the historical max attendance for that group, since
                    the total size of the group is unknown)
                - Students whose name could not be matched. Some students use
                    weird names. This program ignores those students if they
                    are truly unidentifiable, but they get dumped onto the
                    highlight sheet so you can at least see how many anonymous
                    students there are, and what names they are using.
        """
        # master sheet
        master_sheet = self.WB_OUT.create_sheet(title="Master Sheet")
        self.write_master_sheet(master_sheet)
        self.save_workbook(destination)

    def write_master_sheet(self, master_sheet):
        master_sheet.page_setup.fitToWidth = 1
        # header information
        a1 = master_sheet['A1']
        a1.value = 'Zoom Attendance Report'
        a1.font = self.styles['h1']
        b1 = master_sheet['A2']
        b1.value = 'Key'
        b1.font = self.styles['h2']

        # fill in key
        b2, c2 = master_sheet['A3':'B3'][0]
        b2.fill = self.fills['green']
        c2.value = (
            f'Green: Student attended for at least {self.thresholds["green"]} '
            'minutes.'
        )
        b3, c3 = master_sheet['A4':'B4'][0]
        b3.fill = self.fills['yellow']
        c3.value = (
            f'Yellow: Student attended for at least {self.thresholds["yellow"]} '
            'minutes.'
        )
        b4, c4 = master_sheet['A5':'B5'][0]
        b4.fill = self.fills['red']
        c4.value = (
            f'Red: Student attended for at least {self.thresholds["red"]} '
            'minutes.'
        )

        breakpoint()
        if self.meetings.group_dict:
            for meeting_set_name, meetings in self.meetings.group_dict.items():
                self.write_group_dict_item(meeting_name, meetings)

    def write_group_dict_item(
        meeting_name: str,
        meeting: MeetingSet,
        worksheet,
        starting_cell: str
    ):
        """
        Returns the first empty row (int) after the block it wrote.
        """
        # make sure we aren't in a column past z
        try:
            assert len(starting_cell) == 2
            assert starting_cell[1:].isnumerical()
            assert ord(starting_cell[0].upper()) < 90
        except AssertionError:
            raise Exception("This function can not write to columns past Z")
        # get cell range to iterate over
        min_col = starting_cell[0].upper()
        row = int(starting_cell[1:])
        # TODO finish this method
        raise NotImplementedError('Need to finish this method')

    def save_workbook(self, path: Path):
        self.WB_OUT.save(path.resolve())